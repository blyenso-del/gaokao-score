import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import json, os, requests
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

CACHE = r'E:\ai\second\prov_cache'
OUTDIR = r'E:\ai\second\各省录取分数线'
os.makedirs(OUTDIR, exist_ok=True)
YEARS = ['2025', '2024', '2023']
H = {'User-Agent': 'Mozilla/5.0', 'Referer': 'https://www.gaokao.cn/'}

with open(r'E:\ai\second\province_info.json', encoding='utf-8') as f:
    pj = json.load(f)['data']['data']
pid2name = {str(p['value']): p['label'] for p in pj}

def num(x):
    try: return int(x)
    except: return x if x not in (None,'','-') else None

schools = []
for fn in os.listdir(CACHE):
    if not fn.endswith('.json'): continue
    with open(os.path.join(CACHE, fn), encoding='utf-8') as f:
        s = json.load(f)
    if s.get('scores'): schools.append(s)

# 按省份归集: pid -> [row]
by_prov = defaultdict(list)
for s in schools:
    for pid, ys in s['scores'].items():
        row = {'name': s['name'], 'zs_code': s.get('zs_code'), 'level': s.get('level'),
               'home': s.get('home_prov')}
        has = False
        for y in YEARS:
            wu = ys.get(y) or {}
            wl = num(wu.get('2073')); ll = num(wu.get('2074'))
            row[y+'_wl'] = wl; row[y+'_ls'] = ll
            if wl is not None or ll is not None: has = True
        if has: by_prov[pid].append(row)

HDR = ['学校名称','国标码','层次','院校所在地','2025物理类','2025历史类',
       '2024物理类','2024历史类','2023物理类','2023历史类']
KEYS = ['name','zs_code','level','home','2025_wl','2025_ls','2024_wl','2024_ls','2023_wl','2023_ls']
hf = PatternFill('solid', fgColor='2F5496'); ff = Font(color='FFFFFF', bold=True, size=10)
th = Side(style='thin', color='D9D9D9'); bd = Border(left=th,right=th,top=th,bottom=th)

def safe(name): return name.replace('/', '_').replace('\\', '_')

provs = sorted(by_prov.keys(), key=lambda x:int(x) if str(x).isdigit() else 999)
total_rows = 0
made = []
for pid in provs:
    pn = pid2name.get(str(pid), str(pid))
    rows = sorted(by_prov[pid], key=lambda r:-(r.get('2025_wl') or 0))
    total_rows += len(rows)
    base = os.path.join(OUTDIR, safe(pn))

    # Excel
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = safe(pn)[:30]
    ws.append(HDR)
    for ci in range(1, len(HDR)+1):
        c = ws.cell(row=1, column=ci); c.fill=hf; c.font=ff
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border=bd
    for r in rows:
        ws.append([r.get(k) for k in KEYS])
    Wd={1:28,2:9,3:7,4:10}
    for ci in range(1,len(HDR)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = Wd.get(ci,10)
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    wb.save(base + '.xlsx')

    # MD
    def m(v): return '' if v is None else str(v)
    L = [f'# {pn} · 院校录取分数线（2023–2025）\n',
         f'> 数据来源：gaokao.cn · 院校最低分（物理类/历史类）· 共 {len(rows)} 所院校\n',
         '\n| 学校 | 国标码 | 层次 | 2025物理 | 2025历史 | 2024物理 | 2024历史 | 2023物理 | 2023历史 |',
         '|---|---|---|---|---|---|---|---|---|']
    for r in rows:
        L.append('| ' + ' | '.join([m(r['name']), m(r['zs_code']), m(r['level']),
            m(r.get('2025_wl')), m(r.get('2025_ls')), m(r.get('2024_wl')),
            m(r.get('2024_ls')), m(r.get('2023_wl')), m(r.get('2023_ls'))]) + ' |')
    with open(base + '.md', 'w', encoding='utf-8') as f:
        f.write('\n'.join(L))
    made.append((pn, len(rows)))

print(f'生成 {len(made)} 个省份, 各2个文件(MD+Excel), 共 {total_rows} 条记录')
print('输出目录:', OUTDIR)
for pn, n in made:
    print(f'  {pn}: {n} 所')
