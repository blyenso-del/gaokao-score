import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import json, os, re
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

CACHE = r'E:\ai\second\mnzy_cache'
SRC = r'E:\ai\second\2026志愿意向表空白01 的副本.xlsx'
OUT_XLSX = r'E:\ai\second\大学专业历年数据.xlsx'
OUT_MD = r'E:\ai\second\大学专业历年数据.md'
YEARS = ['2025', '2024', '2023']

def num(x):
    if x in (None, '-', '', 'None'): return None
    try: return int(x)
    except:
        try: return float(x)
        except: return x

def parse_triple(s):
    out = {}
    try:
        for item in json.loads(s):
            for yr, val in item.items():
                p = str(val).split(',')
                out[yr] = (p[0] if len(p) > 0 else '-',
                           p[1] if len(p) > 1 else '-',
                           p[2] if len(p) > 2 else '-')
    except: pass
    return out

def parse_single(s):
    out = {}
    try:
        for item in json.loads(s):
            for yr, val in item.items(): out[yr] = val
    except: pass
    return out

wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_src = wb_src['本科']
code2name, order = {}, []
for row in range(2, ws_src.max_row + 1):
    c = ws_src.cell(row=row, column=4).value
    if c is None or not str(c).strip(): continue
    code = str(c).strip().split('.')[0]
    if code not in code2name:
        nm = None
        for off in range(3):
            v = ws_src.cell(row=row + off, column=3).value
            if v: nm = str(v).strip(); break
        code2name[code] = nm or code
        order.append(code)

rows = []
for code in order:
    fp = os.path.join(CACHE, code + '.json')
    if not os.path.exists(fp): continue
    with open(fp, encoding='utf-8') as f:
        data = json.load(f)
    uname = code2name.get(code, code)
    for grp in (data.get('body') or []):
        g = str(grp.get('universityMajorGroup'))
        batchline = {}
        for h in grp.get('histories', []):
            batchline[str(h.get('year'))] = h.get('batchLine')
        for b in grp.get('branches', []):
            bs = parse_triple(b.get('branchScore', ''))
            ss = parse_triple(b.get('schoolScore', ''))
            mp = parse_single(b.get('majorParityScore', ''))
            sp = parse_single(b.get('schoolParityScore', ''))
            row = {
                'univ': uname, 'code': code, 'province': b.get('province'),
                'property': b.get('propertyName'), 'level': b.get('level'),
                'category': b.get('categoryName'), 'batch': b.get('batch'),
                'group': g, 'mcode': b.get('majorCode'),
                'mname': b.get('majorName'),
                'remarks': (b.get('majorRemarks') or '').replace('\n', ' ').strip(),
                'claim': b.get('claim'), 'studyYear': num(b.get('studyYear')),
                'studyCost': num(b.get('studyCost')),
                'planNum': num(b.get('branchPlanNum')),
                'grpPlanNum': num(b.get('schoolPlanNum')),
                'rate': num(b.get('enrolmentRate')),
                'tags': b.get('tags'),
            }
            for yr in YEARS:
                t = bs.get(yr, ('-', '-', '-'))
                row[yr + '_mScore'] = num(t[0])
                row[yr + '_mPlace'] = num(t[1])
                row[yr + '_mEnroll'] = num(t[2])
                row[yr + '_mParity'] = num(mp.get(yr))
            for yr in YEARS:
                t = ss.get(yr, ('-', '-', '-'))
                row[yr + '_gScore'] = num(t[0])
                row[yr + '_gPlace'] = num(t[1])
                row[yr + '_gParity'] = num(sp.get(yr))
                row[yr + '_batchLine'] = num(batchline.get(yr))
            rows.append(row)

print('共', len(rows), '个专业行,', len(order), '所大学')

order_idx = {c: i for i, c in enumerate(order)}
rows.sort(key=lambda r: (order_idx.get(r['code'], 999), num(r['group']) or 0, num(r['mcode']) or 0))

# 中文表头映射
HDR = {
    'univ': '大学名称', 'code': '院校代码', 'province': '省份', 'property': '办学性质',
    'level': '层次', 'category': '招生类别', 'batch': '批次', 'group': '专业组',
    'mcode': '专业代码', 'mname': '专业名称', 'remarks': '专业备注', 'claim': '选科要求',
    'studyYear': '学制(年)', 'studyCost': '学费(元/年)', 'planNum': '专业计划数',
    'grpPlanNum': '专业组计划数', 'rate': '录取率(%)', 'tags': '院校标签',
}
for yr in YEARS:
    HDR[yr + '_mScore'] = yr + '专业最低分'
    HDR[yr + '_mPlace'] = yr + '专业最低位次'
    HDR[yr + '_mEnroll'] = yr + '专业录取数'
    HDR[yr + '_mParity'] = yr + '专业投档分'
for yr in YEARS:
    HDR[yr + '_gScore'] = yr + '专业组最低分'
    HDR[yr + '_gPlace'] = yr + '专业组最低位次'
    HDR[yr + '_gParity'] = yr + '专业组投档分'
    HDR[yr + '_batchLine'] = yr + '批次线'

keys = list(HDR.keys())

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '专业历年数据'
hdr_fill = PatternFill('solid', fgColor='2F5496')
hdr_font = Font(color='FFFFFF', bold=True, size=10)
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
ws.append([HDR[k] for k in keys])
for ci in range(1, len(keys) + 1):
    cell = ws.cell(row=1, column=ci)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
for r in rows:
    ws.append([r.get(k) for k in keys])
widths = {'univ': 22, 'mname': 26, 'remarks': 50, 'tags': 30, 'claim': 9, 'category': 9}
for ci, k in enumerate(keys, 1):
    L = openpyxl.utils.get_column_letter(ci)
    ws.column_dimensions[L].width = widths.get(k, 11)
ws.freeze_panes = 'C2'
ws.auto_filter.ref = ws.dimensions
wb.save(OUT_XLSX)
print('已写 Excel:', OUT_XLSX)

def md(v):
    return '' if v is None else str(v).replace('|', '\\|')

lines = []
lines.append('# 大学专业历年录取数据（2023–2025）\n')
lines.append('> 数据来源：mnzy.gaokao.cn · 广东 物理类 · 共 ' + str(len(order)) + ' 所大学、' + str(len(rows)) + ' 个专业\n')
lines.append('> 每格为「最低分 / 最低位次 / 录取数」；批次线为该专业组所在批次省控线。\n')
lines.append('\n## 目录\n')
for c in order:
    lines.append('- ' + code2name[c] + '（' + c + '）')
lines.append('')

by_univ = defaultdict(list)
for r in rows: by_univ[r['code']].append(r)

for c in order:
    urows = by_univ.get(c, [])
    if not urows: continue
    nm = code2name[c]
    tags = urows[0].get('tags') or ''
    lines.append('\n## ' + nm + '（' + c + '）\n')
    if tags: lines.append('**标签**：' + tags + '\n')
    lines.append('专业数：' + str(len(urows)) + '\n')
    lines.append('\n| 组 | 代码 | 专业名称 | 选科 | 学制 | 学费 | 计划 | 2025 分/位次/录取 | 2024 分/位次/录取 | 2023 分/位次/录取 | 批次线 25/24/23 |')
    lines.append('|---|---|---|---|---|---|---|---|---|---|---|')
    for r in urows:
        def yc(yr):
            s, p, n = r.get(yr + '_mScore'), r.get(yr + '_mPlace'), r.get(yr + '_mEnroll')
            if s is None and p is None: return '—'
            return md(s) + '/' + md(p) + '/' + md(n)
        bl = '/'.join(md(r.get(yr + '_batchLine')) for yr in YEARS)
        name = md(r['mname'])
        rem = r.get('remarks') or ''
        if rem: name += '<br><sub>' + md(rem[:60]) + '</sub>'
        cells = [md(r['group']), md(r['mcode']), name, md(r['claim']),
                 md(r['studyYear']), md(r['studyCost']), md(r['planNum']),
                 yc('2025'), yc('2024'), yc('2023'), bl]
        lines.append('| ' + ' | '.join(cells) + ' |')
    lines.append('')

with open(OUT_MD, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print('已写 MD:', OUT_MD)
