import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import json, re, os, shutil
import openpyxl
from openpyxl.utils import get_column_letter

SRC = r'E:\ai\second\2026志愿意向表空白01 的副本.xlsx'
BAK = r'E:\ai\second\2026志愿意向表空白01 的副本_备份2.xlsx'
CACHE = r'E:\ai\second\mnzy_cache'

def code_int(s):
    try: return int(re.sub(r'\D','',str(s)))
    except: return None

def build(data):
    by_code, by_grp_code, grp_branches = {}, {}, {}
    for grp in (data.get('body') or []):
        g = str(grp.get('universityMajorGroup'))
        for b in grp.get('branches', []):
            ci = code_int(b.get('majorCode'))
            yp = {}
            try:
                for item in json.loads(b['branchScore']):
                    for yr,val in item.items():
                        parts = val.split(','); yp[yr] = parts[1] if len(parts)>=2 else '-'
            except: pass
            rec = {'group':g,'name':b.get('majorName'),'years':yp,'remarks':b.get('majorRemarks') or ''}
            by_code.setdefault(ci, []).append(rec)
            by_grp_code[(g, ci)] = rec
            grp_branches.setdefault(g, []).append(rec)
    return by_code, by_grp_code, grp_branches

def ex_code(text):
    if text is None: return None
    m = re.search(r'代码[:：]\s*([0-9A-Za-z]+)', str(text))
    return m.group(1) if m else None

def match(maps, g, ci, text):
    by_code, by_grp_code, grp_branches = maps
    if ci is not None:
        m = by_grp_code.get((g, ci))
        if m: return m
        lst = by_code.get(ci)
        if lst:
            if len(lst)==1: return lst[0]
            for r in lst:
                if r['group']==g: return r
    # 名称回退(无代码)
    core = re.split(r'[（(]', str(text))[0].strip()
    cands = grp_branches.get(g, [])
    hits = [r for r in cands if r['name']==core]
    if len(hits)==1: return hits[0]
    if len(hits)>1:
        # 用括号内备注消歧
        paren = re.findall(r'[（(]([^（）()]+)[）)]', str(text))
        kw = paren[0] if paren else ''
        for r in hits:
            if kw and (kw in r['remarks'] or kw in (r['name'] or '')):
                return r
    return None

# ---- 缓存载入 ----
cache = {}
for fn in os.listdir(CACHE):
    if fn.endswith('.json'):
        with open(os.path.join(CACHE, fn), encoding='utf-8') as f:
            cache[fn[:-5]] = build(json.load(f))

# ---- 备份 + 打开(可写) ----
shutil.copyfile(SRC, BAK)
wb = openpyxl.load_workbook(SRC)
ws = wb['本科']
name_cols = [8,11,14,17,20,23]

records = []
for row in range(2, ws.max_row+1):
    c = ws.cell(row=row, column=4).value
    if c is not None and str(c).strip():
        records.append((row, str(c).strip().split('.')[0]))

total=ok=fail=filled=0
fails=[]
for rec_row, ucode in records:
    gv = ws.cell(row=rec_row, column=7).value
    g = str(gv).split('.')[0] if gv is not None else None
    maps = cache.get(ucode)
    if not maps:
        continue
    for nc in name_cols:
        nv = ws.cell(row=rec_row, column=nc).value
        if nv is None or not str(nv).strip(): continue
        total += 1
        m = match(maps, g, code_int(ex_code(nv)), nv)
        if not m:
            fail += 1
            fails.append((rec_row, ucode, g, str(nv).replace('\n','|')[:32]))
            continue
        ok += 1
        rank_col = nc + 2
        for off in range(3):
            yr = re.search(r'(20\d{2})', str(ws.cell(row=rec_row+off, column=nc+1).value or ''))
            if not yr: continue
            place = m['years'].get(yr.group(1), '-')
            if place and place != '-':
                ws.cell(row=rec_row+off, column=rank_col).value = int(re.sub(r'\D','',place))
                filled += 1

wb.save(SRC)
print(f"专业总数={total} 成功={ok} 失败={fail} 填入排位={filled}")
if fails:
    print("失败:")
    for f in fails: print(f"  行{f[0]} 院校{f[1]} 组{f[2]} '{f[3]}'")
print(f"备份: {BAK}")
print(f"已保存: {SRC}")
