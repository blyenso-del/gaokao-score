import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import json, os, re
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

CACHE = r'E:\ai\second\mnzy_cache_all'
OUT_XLSX = r'E:\ai\second\全国高校广东录取数据.xlsx'
OUT_MD = r'E:\ai\second\全国高校广东录取数据.md'
YEARS = ['2025', '2024', '2023']

def num(x):
    if x in (None, '-', '', 'None'): return None
    try: return int(x)
    except:
        try: return float(x)
        except: return x

def parse_triple(s):
    out = {}
    try:
        for item in json.loads(s):
            for yr, val in item.items():
                p = str(val).split(',')
                out[yr] = (p[0] if len(p) > 0 else '-',
                           p[1] if len(p) > 1 else '-',
                           p[2] if len(p) > 2 else '-')
    except: pass
    return out

def parse_single(s):
    out = {}
    try:
        for item in json.loads(s):
            for yr, val in item.items(): out[yr] = val
    except: pass
    return out

def expand(body, uname, code, level, classify):
    rows = []
    for grp in (body or []):
        g = str(grp.get('universityMajorGroup'))
        batchline = {}
        for h in grp.get('histories', []):
            batchline[str(h.get('year'))] = h.get('batchLine')
        for b in grp.get('branches', []):
            bs = parse_triple(b.get('branchScore', ''))
            ss = parse_triple(b.get('schoolScore', ''))
            mp = parse_single(b.get('majorParityScore', ''))
            sp = parse_single(b.get('schoolParityScore', ''))
            row = {
                'univ': uname, 'code': code, 'level': level, 'classify': classify,
                'province': b.get('province'), 'property': b.get('propertyName'),
                'category': b.get('categoryName'), 'batch': b.get('batch'),
                'group': g, 'mcode': b.get('majorCode'), 'mname': b.get('majorName'),
                'remarks': (b.get('majorRemarks') or '').replace('\n', ' ').strip(),
                'claim': b.get('claim'), 'studyYear': num(b.get('studyYear')),
                'studyCost': num(b.get('studyCost')), 'planNum': num(b.get('branchPlanNum')),
                'grpPlanNum': num(b.get('schoolPlanNum')), 'rate': num(b.get('enrolmentRate')),
                'tags': b.get('tags'),
            }
            for yr in YEARS:
                t = bs.get(yr, ('-', '-', '-'))
                row[yr+'_mScore'] = num(t[0]); row[yr+'_mPlace'] = num(t[1])
                row[yr+'_mEnroll'] = num(t[2]); row[yr+'_mParity'] = num(mp.get(yr))
            for yr in YEARS:
                t = ss.get(yr, ('-', '-', '-'))
                row[yr+'_gScore'] = num(t[0]); row[yr+'_gPlace'] = num(t[1])
                row[yr+'_gParity'] = num(sp.get(yr)); row[yr+'_batchLine'] = num(batchline.get(yr))
            rows.append(row)
    return rows

files = [f for f in os.listdir(CACHE) if f.endswith('.json')]
schools = []
for fn in files:
    with open(os.path.join(CACHE, fn), encoding='utf-8') as f:
        schools.append(json.load(f))
schools.sort(key=lambda s: int(s['code']))

# 用 zsgkId -> (schoolid.json + school_code.json) 反查校名(比 logo 更准更全)
with open(r'E:\ai\second\school_code.json', encoding='utf-8') as f:
    sid2name = {str(v['school_id']): v['name'] for v in json.load(f)['data'].values()}
with open(r'E:\ai\second\schoolid.json', encoding='utf-8') as f:
    sid2name.update({str(x['school_id']): x['name'] for x in json.load(f)['data']['school']})
def zsgk_name(s):
    for body in (s.get('phys'), s.get('hist')):
        for g in (body or []):
            for b in g.get('branches', []):
                z = b.get('zsgkId')
                if z is not None and str(z) in sid2name:
                    return sid2name[str(z)]
    return None
for s in schools:
    nm = zsgk_name(s)
    if nm: s['name'] = nm

all_rows = []
unknown = []
for s in schools:
    if s['name'].startswith('院校'): unknown.append(s['code'])
    all_rows += expand(s.get('phys'), s['name'], s['code'], s['level'], '物理')
    all_rows += expand(s.get('hist'), s['name'], s['code'], s['level'], '历史')

n_ben = sum(1 for s in schools if s['level'] == '本科')
n_zhuan = sum(1 for s in schools if s['level'] == '专科')
print('院校', len(schools), '(本科', n_ben, '/专科', n_zhuan, '), 专业行', len(all_rows), ', 未识别校名', len(unknown))

HDR = {'univ':'大学名称','code':'国标码','level':'层次','classify':'科类','province':'省份',
       'property':'办学性质','category':'招生类别','batch':'批次','group':'专业组','mcode':'专业代码',
       'mname':'专业名称','remarks':'专业备注','claim':'选科要求','studyYear':'学制(年)',
       'studyCost':'学费(元/年)','planNum':'专业计划数','grpPlanNum':'专业组计划数','rate':'录取率(%)','tags':'院校标签'}
for yr in YEARS:
    HDR[yr+'_mScore']=yr+'专业最低分'; HDR[yr+'_mPlace']=yr+'专业最低位次'
    HDR[yr+'_mEnroll']=yr+'专业录取数'; HDR[yr+'_mParity']=yr+'专业投档分'
for yr in YEARS:
    HDR[yr+'_gScore']=yr+'专业组最低分'; HDR[yr+'_gPlace']=yr+'专业组最低位次'
    HDR[yr+'_gParity']=yr+'专业组投档分'; HDR[yr+'_batchLine']=yr+'批次线'
keys = list(HDR.keys())

wb = openpyxl.Workbook()
ws = wb.active; ws.title = '全国专业历年数据'
hf = PatternFill('solid', fgColor='2F5496'); ff = Font(color='FFFFFF', bold=True, size=10)
th = Side(style='thin', color='D9D9D9'); bd = Border(left=th, right=th, top=th, bottom=th)
ws.append([HDR[k] for k in keys])
for ci in range(1, len(keys)+1):
    c = ws.cell(row=1, column=ci); c.fill=hf; c.font=ff
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); c.border=bd
for r in all_rows:
    ws.append([r.get(k) for k in keys])
widths = {'univ':24,'mname':26,'remarks':46,'tags':28,'claim':9,'category':9,'classify':6,'level':6}
for ci, k in enumerate(keys, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = widths.get(k, 11)
ws.freeze_panes = 'E2'; ws.auto_filter.ref = ws.dimensions
wb.save(OUT_XLSX)
print('已写 Excel:', OUT_XLSX)

def md(v): return '' if v is None else str(v).replace('|', '\\|')

by_code = defaultdict(list)
for r in all_rows: by_code[r['code']].append(r)

L = []
L.append('# 全国高校广东录取数据（2023–2025）\n')
L.append('> 数据来源：mnzy.gaokao.cn · 广东 · 物理类(物理化生)+历史类(历史政地)\n')
L.append('> 共 **' + str(len(schools)) + '** 所院校（本科 ' + str(n_ben) + ' / 专科 ' + str(n_zhuan) + '）、**' + str(len(all_rows)) + '** 条专业记录\n')
L.append('> 每格为「最低分 / 最低位次 / 录取数」；批次线为该专业组所在批次省控线。\n')
L.append('\n## 目录\n')
for lv in ['本科', '专科']:
    sl = [s for s in schools if s['level'] == lv]
    L.append('\n### ' + lv + '（' + str(len(sl)) + ' 所）\n')
    line = ['[' + s['name'] + '](#' + s['code'] + ')（' + s['code'] + '）' for s in sl]
    for i in range(0, len(line), 5):
        L.append(' · '.join(line[i:i+5]))
L.append('')

def render_table(rows):
    out = ['| 组 | 代码 | 专业名称 | 选科 | 学制 | 学费 | 计划 | 2025 分/位次/录取 | 2024 | 2023 | 批次线25/24/23 |',
           '|---|---|---|---|---|---|---|---|---|---|---|']
    for r in rows:
        def yc(yr):
            s, p, n = r.get(yr+'_mScore'), r.get(yr+'_mPlace'), r.get(yr+'_mEnroll')
            if s is None and p is None: return '—'
            return md(s) + '/' + md(p) + '/' + md(n)
        bl = '/'.join(md(r.get(yr+'_batchLine')) for yr in YEARS)
        name = md(r['mname'])
        rem = r.get('remarks') or ''
        if rem: name += '<br><sub>' + md(rem[:50]) + '</sub>'
        out.append('| ' + ' | '.join([md(r['group']), md(r['mcode']), name, md(r['claim']),
                    md(r['studyYear']), md(r['studyCost']), md(r['planNum']),
                    yc('2025'), yc('2024'), yc('2023'), bl]) + ' |')
    return out

for s in schools:
    code = s['code']
    rows = by_code.get(code, [])
    if not rows: continue
    tags = rows[0].get('tags') or ''
    L.append('\n<a id="' + code + '"></a>\n## ' + s['name'] + '（' + code + '） [' + s['level'] + ']\n')
    if tags: L.append('**标签**：' + tags + '\n')
    phys = [r for r in rows if r['classify'] == '物理']
    hist = [r for r in rows if r['classify'] == '历史']
    if phys:
        L.append('\n**物理类**\n'); L += render_table(phys)
    if hist:
        L.append('\n**历史类**\n'); L += render_table(hist)
    L.append('')

with open(OUT_MD, 'w', encoding='utf-8') as f:
    f.write('\n'.join(L))
sz = os.path.getsize(OUT_MD) / 1024 / 1024
print('已写 MD:', OUT_MD, '(', round(sz, 1), 'MB)')
if unknown:
    print('未识别校名的码(', len(unknown), '):', unknown[:20])
